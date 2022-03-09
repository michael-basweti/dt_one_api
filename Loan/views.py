from functools import partial
from rest_framework.response import Response
from rest_framework.views import APIView
from datetime import date
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse

from .models import Loans, Payavenue, Vwloans, Vwunpaidloans
from .serializers import LoansSerializer, PaymentsSerializer, PayavenueSerializer, VwLoansSerializer, VwUnpaidLoansSerializer
from rest_framework import permissions, status
from utils.register_email import loan_acknowledgement,loan_approval, loan_denied, reminder_email
from openpyxl import Workbook

class LoansView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def post(self, request):
        serializer = LoansSerializer(data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save(requestedby=request.user)
            emailUser = request.user.email
            # aws_email_send(emailUser, password)
            # gmail_send_email(emailUser, password)
            loan_acknowledgement(emailUser)
            data = {
                "data": serializer.data,
                "message": "Loan successfully submitted"
            }
            return Response(data, status=status.HTTP_201_CREATED)
        else:
            return Response((serializer.errors), status=status.HTTP_400_BAD_REQUEST)


class ApproveView(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def post(self, request):
        loanid = request.data.get('loanid')
        amountdispatched = request.data.get('amountdispatched')
        approvalcomments = request.data.get('approvalcomments')

        try:
            loan = Loans.objects.get(loanid=loanid)
        except Loans.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

        dataLoan = {
            "amountdispatched":amountdispatched,
            "approvalcomments":approvalcomments,
            "approved":True,
            "remainingamount":amountdispatched
        }

        datapay = {
            "amountpaid":amountdispatched,
            "paymenttype":2,
            "payavenue":loan.payavenue.payavenueid
        }

        loanSerializer = LoansSerializer(loan,data=dataLoan,partial=True)
        payserializer = PaymentsSerializer(data=datapay,partial=True)

        if loanSerializer.is_valid():
            if payserializer.is_valid():
                loanSerializer.save(approvedby=request.user)
                payserializer.save(processedby=request.user)

                emailUser = loan.requestedby.email

                loan_approval(emailUser)

                data = {
                    "message":"Loan Approved"
                }

                return Response(data, status=status.HTTP_200_OK)
            else:
                return Response(payserializer.errors, status=status.HTTP_200_OK)
        else:
            return Response(loanSerializer.errors, status=status.HTTP_200_OK)



class DenyLoan(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def post(self,request):
        loanid = request.data.get('loanid')
        deniedreason = request.data.get('deniedreason')


        data = {
            "deniedreason":deniedreason,
            "denied":True
        }

        try:
            loan = Loans.objects.get(loanid=loanid)
        except Loans.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

        loanSerializer = LoansSerializer(loan,data=data,partial=True)

        if loanSerializer.is_valid():
            loanSerializer.save(approvedby=request.user)
            emailUser = loan.requestedby.email
            loan_denied(emailUser)
            data = {
                "message":"Loan denied"
            }

            return Response(data, status=status.HTTP_200_OK)

        else:
            return Response(loanSerializer.errors, status=status.HTTP_200_OK)



# Get Pay Avenues

class GetPayAvenues(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request):
        avenues = Payavenue.objects.all()
        serializer = PayavenueSerializer(avenues, many=True)
        return Response(serializer.data)


class GetUnProcessedLoans(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request):
        loans = Vwloans.objects.all().exclude(approved=True).exclude(denied=True)
        serializer = VwLoansSerializer(loans, many=True)
        return Response(serializer.data)


class GetOneUnProcessedLoans(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request,loanid):
        loans = Vwloans.objects.get(loanid=loanid)
        serializer = VwLoansSerializer(loans)
        return Response(serializer.data)

class GetUserLoans(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request):
        loans = Vwloans.objects.filter(requestedby=request.user.id)
        serializer = VwLoansSerializer(loans, many=True)
        return Response(serializer.data)


class GetLoansDue(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request, startdate, enddate):
        today = date.today().strftime("%Y-%m-%d")
        if(startdate==today and enddate==today):
            loans = Vwunpaidloans.objects.all()
            serializer = VwUnpaidLoansSerializer(loans, many=True)
            return Response(serializer.data)
        else:
            loans = Vwunpaidloans.objects.filter(paymentdate__lte=enddate,paymentdate__gte=startdate)
            serializer = VwUnpaidLoansSerializer(loans, many=True)
            return Response(serializer.data)



class CreatePayment(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def post(self,request):
        loanid = request.data.get('loanid')
        amountpaid = request.data.get('amount')

        try:
            loan = Loans.objects.get(loanid=loanid)
        except Loans.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

        
        approved = loan.approved

        if approved:
            amountremaining = float(loan.remainingamount)
            amounttobepaid = float(amountpaid)
            amountdifference = amountremaining - amounttobepaid

            if amountdifference < 0:
                postiveamountdifference = amountdifference*(-1)
                data = {
                    "message":"You are overpaying by {}".format(postiveamountdifference)
                }

                return Response(data=data,status=status.HTTP_404_NOT_FOUND)
            else:
                if amountdifference > 0:
                    fullypaid = False
                elif amountdifference == 0:
                    fullypaid = True
                dataLoan = {
                    "fullypaid":fullypaid,
                    "remainingamount":amountdifference
                }

                datapay = {
                    "amountpaid":amountpaid,
                    "paymenttype":1,
                    "payavenue":loan.payavenue.payavenueid,
                    "loanid":loanid
                }

                loanSerializer = LoansSerializer(loan,data=dataLoan,partial=True)
                payserializer = PaymentsSerializer(data=datapay,partial=True)
                if loanSerializer.is_valid():
                    if payserializer.is_valid():
                        loanSerializer.save(approvedby=request.user)
                        payserializer.save(processedby=request.user)

                        # emailUser = loan.requestedby.email

                        # loan_approval(emailUser)

                        data = {
                            "message":"Payment Processed"
                        }

                        return Response(data, status=status.HTTP_200_OK)
                    else:
                        return Response(payserializer.errors, status=status.HTTP_200_OK)
                else:
                    return Response(loanSerializer.errors, status=status.HTTP_200_OK)
        else:
            data = {
                'message':'You cannot pay for an undispatched loan'
            }
            return Response(data=data,status=status.HTTP_404_NOT_FOUND)


@api_view(['GET',])
@permission_classes((IsAuthenticated, ))
def DueLoansExcel(request, startdate,enddate):
    today = date.today().strftime("%Y-%m-%d")
    if(startdate==today and enddate==today):
        order_queryset = Vwunpaidloans.objects.all()
    else:
        order_queryset = Vwunpaidloans.objects.filter(paymentdate__lte=enddate,paymentdate__gte=startdate)

    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    response['Content-Disposition'] = 'attachment; filename=loansdue.xlsx'
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Loans Due'

    # Define the titles for columns
    columns = [
        'LoanId',
        'Requested By',
        'Pay Through',
        'Due Date',
        'Amount Dispatched',
        'Amount Paid',
        'Amount Remaining',
        'Email',
        'Phone'
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all orders
    for order in order_queryset:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            order.loanid,
            order.requestedname,
            order.payavenuedescription,
            order.paymentdate,
            order.amountdispatched,
            order.amountpaid,
            order.remainingamount,
            order.email,
            order.phone
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


class SendReminderEmails(APIView):
    permission_classes = (permissions.IsAuthenticated,)

    def get(self, request, startdate, enddate):
        today = date.today().strftime("%Y-%m-%d")
        if(startdate==today and enddate==today):
            loans = Vwunpaidloans.objects.all()
        else:
            loans = Vwunpaidloans.objects.filter(paymentdate__lte=enddate,paymentdate__gte=startdate)
        
        emails = []

        for loan in loans:
            emails.append(loan.email)

        print(emails)

        new_lst=(','.join(emails))
        reminder_email(emails)
        data = {
            "message": "Emails successfully sent"
        }
        return Response(data, status=status.HTTP_201_CREATED)